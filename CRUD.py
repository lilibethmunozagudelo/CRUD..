from datetime import datetime
from openpyxl import load_workbook

ruta=r"BASECRUD.xlsx" 

def leer(ruta:str, extraer:str):
    archivo_exccel=load_workbook(ruta)
    hoja_datos=archivo_exccel["DATOS DEL CRUD"]
    hoja_datos=hoja_datos["A2":"F"+str(hoja_datos.max_row)]

    info={}

    for i in hoja_datos:
        if isinstance(i[0].value,int):
            info.setdefault(i[0].value,{"tarea":i[1].value,"descripcion":i[2].value,"estado":i[3].value,"fecha_inicio":i[4].value,"fecha_finalizacion":i[5].value})

    if not(extraer=="todo"):
        info=filtrar(info,extraer)

    for i in info:
        print("*** Tarea ****")
        print("id: "+str(i)+"\n"+"titulo: "+str(info[i]["tarea"])+"\n"+"descripcion: "+str(info[i]["descripcion"])+"\n"+"estado: "+str(info[i]["estado"])+"\n"+"fecha creacion: "+str(info[i]["fecha_inicio"])+"\n"+"fecha de finalizacion: "+str(info[i]["fecha_finalizacion"]))
        print()

    return

def filtrar(info:dict,filtro:str):
    aux={}

    for i in info:
        if info[i]["estado"]==filtro:
            aux.setdefault(i,info[i])
    return aux

def actualizar(ruta:str, identificador:int,datos_actualizados:dict):
    archivo_exccel= load_workbook(ruta)
    hoja_datos=archivo_exccel["DATOS DEL CRUD"]
    hoja_datos=hoja_datos["A2":"F"+str(hoja_datos.max_row)]
    hoja=archivo_exccel.active

    titulo=2
    descripcion=3
    estado=4
    fecha_inicio=5
    fecha_finalizacion=6
    encontro=False
    for i in hoja_datos:
        if i[0].value==identificador:
            fila=i[0].row
            encontro=True
            for d in datos_actualizados:
                if d=="titulo" and not(datosActualizados[d]==""):
                    hoja.cell(row=fila, column=titulo).value=datosActualizados[d]
                elif d=="descripcion" and not(datosActualizados[d]==""):
                    hoja.cell(row=fila, column=descripcion).value=datosActualizados[d]
                elif d=="estado" and not(datosActualizados[d]==""):
                    hoja.cell(row=fila, column=estado).value=datosActualizados[d]
                elif d=="fecha_inicio" and not(datosActualizados[d]==""):
                    hoja.cell(row=fila, column=fecha_inicio).value=datosActualizados[d]
                elif d=="fecha_finalizacion" and not(datosActualizados[d]==""):
                    hoja.cell(row=fila, column=fecha_finalizacion).value=datosActualizados[d]
    archivo_exccel.save(ruta)
    if encontro==False:
        print("Error: No existe una tarea con es Id")
        print()
    return

def agregar(ruta:int, datos:dict):
    archivo_exccel=load_workbook(ruta)
    hoja_datos=archivo_exccel["DATOS DEL CRUD"]
    hoja_datos=hoja_datos["A2":"F"+str(hoja_datos.max_row+1)]
    hoja=archivo_exccel.active

    titulo=2
    descripcion=3
    estado=4
    fecha_inicio=5
    fecha_finalizacion=6
    for i in hoja_datos:

        if not( isinstance(i[0].value, int)):
            identificador=i[0].row
            hoja.cell(row=identificador, column=1).value=identificador-1
            hoja.cell(row=identificador, column=titulo).value=datos["titulo"]
            hoja.cell(row=identificador, column=descripcion).value=datos["descripcion"]
            hoja.cell(row=identificador, column=estado).value=datos["estado"]
            hoja.cell(row=identificador, column=fecha_inicio).value=datos["fecha_inicio"]
            hoja.cell(row=identificador, column=fecha_finalizacion).value=datos["fecha_finalizacion"]
            break
    archivo_exccel.save(ruta)
    return

def borrar(ruta,identificador):
    archivo_exccel=load_workbook(ruta)
    hoja_datos=archivo_exccel["DATOS DEL CRUD"]
    hoja_datos=hoja_datos["A2":"F"+str(hoja_datos.max_row)]
    hoja=archivo_exccel.active 

    titulo=2
    descripcion=3
    estado=4
    fecha_inicio=5
    fecha_finalizacion=6
    encontro=False
    for i in hoja_datos:
        if i[0].value==identificador:
            fila=i[0].row
            encontro=True

            hoja.cell(row=fila, column=1).value=""
            hoja.cell(row=fila, column=titulo).value=""
            hoja.cell(row=fila, column=descripcion).value=""
            hoja.cell(row=fila, column=estado).value=""
            hoja.cell(row=fila, column=fecha_inicio).value=""
            hoja.cell(row=fila, column=fecha_finalizacion).value=""
    archivo_exccel.save(ruta)
    if encontro==False:
        print("Error: No existe una tarea con ese Id")
        print()
    return



datosActualizados={"titulo":"", "descripcion":"", "estado":"", "fecha_inicio":"", "fecha_finalizacion":""}
while True:
    print("Indique la accion que desea realizar: ")
    print("Consultar: 1")
    print("Actualizar: 2")
    print("Crear nueva tarea: 3")
    print("Borrar: 4")
    accion= input("Escriba la opcion: ")
    if not(accion=="1") and not(accion=="2") and not(accion=="3") and not(accion=="4"):
        print("Comando invalido por favor elija una opcion valida")
    elif accion=="1":
        opc_consulta=""
        print("Indique la tarea que desea consultar: ")
        print("Todas las tareas: 1")
        print("En espera: 2")
        print("En ejecucion: 3")
        print("Por aprobar: 4")
        print("Finalizada: 5")
        opc_consulta=input("Escriba la tarea que desea consutar: ")
        if opc_consulta=="1":
            print()
            print()
            print("* Consultando todas las tareas *")
            leer(ruta, "todo")
        elif opc_consulta=="2":
            print()
            print()
            print("* Consultando tareas en espera *")
            leer(ruta, "En espera")
        elif opc_consulta=="3":
            print()
            print()
            print("* Consultando tareas en ejecucion *")
            leer(ruta, "En ejecucion")
        elif opc_consulta=="4":
            print()
            print()
            print("* Consultando tareas por aprobar *")
            leer(ruta, "Por aprobar")
        elif opc_consulta=="5":
            print()
            print()
            print("* Consultando tareas finalizadas *")
            leer(ruta,"Finalizada")
    elif accion=="2":
        datosActualizados={"titulo":"", "descripcion":"", "estado":"", "fecha_inicio":"", "fecha_finalizacion":""}
        print("** Actualizar tarea **")
        print()
        id_actulizar=int(input("Indique el Id de la tarea que desea actualizar: "))
        print()
        print("* Nuevo titulo *")
        print("* Nota: si no desea actualizar el titulo solo oprima ENTER *")
        datosActualizados["titulo"]=input("Indique el nuevo titulo de la tarea: ")
        print()
        print("* Nueva descripcion *")
        print("* Nota: si no desea actualizar la descripcion solo oprima ENTER *")
        datosActualizados["descripcion"]=input("Indique la nueva descripcion de la tarea: ")
        print()
        print("* Nuevo estado *")
        print("En espera: 2")
        print("En ejecucion: 3")
        print("Por aprobar: 4")
        print("Finalizada: 5")
        print("* Nota: si no desea actualizar el estado solo oprima ENTER *")
        estadonuevo=input("Indique el nuevo estado de la tarea: ")
        if estadonuevo=="2":
            datosActualizados["estado"]="En espera"
        elif estadonuevo=="3":
            datosActualizados["estado"]="En ejecucion"
        elif estadonuevo=="4":
            datosActualizados["estado"]="Por aprobar"
        elif estadonuevo=="5":
            now= datetime.now()
            datosActualizados["estado"]="Finalizada"
            datosActualizados["fecha_finalizacion"]=str(now.day) +"/"+ str(now.month) +"/"+ str(now.year)
            now = datetime.now()
            datosActualizados["fecha_inicio"]=str(now.day) +"/"+ str(now.month) +"/"+ str(now.year)
            actualizar(ruta, id_actulizar, datosActualizados)
            print()
    elif accion=="3":
        datosActualizados={"tarea":"", "descripcion":"", "estado":"", "fecha_inicio":"", "fecha_finalizacion":""}
        print("* Crear nueva tarea *")
        print()
        print("* titulo *")
        print()
        datosActualizados["titulo"]=input("Indique el titulo de la tarea: ")
        print()
        print("* descripcion *")
        datosActualizados["descripcion"]=input("Indique la descripcion de la tarea: ")
        print()
        datosActualizados["estado"]="En espera"
        now = datetime.now()
        datosActualizados["fecha_inicio"]=str(now.day) +"/"+ str(now.month) +"/"+ str(now.year)
        datosActualizados["fecha_finalizacion"]=""
        agregar(ruta,datosActualizados)
    elif accion=="4":
        print("")
        print("* Eliminar tarea *")
        iden=int(input("Indique el Id de la tarea que desea eliminar: "))
        borrar(ruta,iden)